import time
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from pynput.keyboard import Key, Controller
from seleniumbase import BaseCase

filepath = "C:/Users/Raghav/Desktop/contact_list.xlsx"

wb = load_workbook(filepath)
sheet = wb.active
frequency = 5
i = 1
# open chrome
driver = webdriver.Chrome('C:\chromedriver')
# open whatsapp homepage for scanning
driver.get("https://web.whatsapp.com")
# wait for some time(in sec)
time.sleep(20)
# open new window
driver.execute_script("window.open('');")
keyboard = Controller()


def shortcut():
    driver.find_element_by_xpath('//*[@title="Type a message"]').send_keys(Keys.SHIFT, Keys.ENTER)


for i in range(frequency):
    b3 = sheet.cell(row=i + 1, column=2)
    url = "https://wa.me/+91"
    # prints out the link
    print((b3.value))
    #

    print(url + str(b3.value))
    # Switch to the new window
    driver.switch_to.window(driver.window_handles[1])
    # open the link
    driver.get(url + str(b3.value))
    # wait for some time(in sec)
    time.sleep(2)
    # locating continue to chat
    continue_to_chat = driver.find_element_by_id("action-button")
    # continue_to_chat.click()
    continue_to_chat.click()
    time.sleep(1)
    # clicking on use whatsapp web
    use_whatsapp_web = driver.find_element_by_xpath("// a[contains(text(),'use WhatsApp Web')]")
    use_whatsapp_web.click()
    # waiting for the no. to load
    time.sleep(25)
    # attaching image
    driver.find_element_by_xpath('//div[@title = "Attach"]').click()
    image_box = driver.find_element_by_xpath(
        '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')
    # locating image
    image_box.send_keys('C:/Users/Raghav/Desktop/Brochure.jpg')
    time.sleep(3)
    # pressing send button
    send_button = driver.find_element_by_xpath('//span[@data-icon="send"]')
    send_button.click()
    # wait for some time
    time.sleep(3)
    # typing message
    message = driver.find_element_by_xpath('//*[@title="Type a message"]')
    message.send_keys("*“SUMMER CAMP 2022”* @ Meera Marg, Bani Park ")
    shortcut()
    message.send_keys("Enhance your life skills learn something new under Professional guidance  - "
                      "Musical Instruments, Skating, Dance, Drawing, Painting, Art & Craft, Spoken English, "
                      "Handwriting, UCMAS, Fitness for Kids, Tuition PG to VIII ")
    shortcut()
    message.send_keys("SPECIAL PACKAGES FOR KIDS ")
    shortcut()
    message.send_keys("AIR-COOLED CAMPUS, TRANSPORT AVAILABLE")
    shortcut()
    shortcut()
    message.send_keys("Venue: Modern Gurukul Academy ")
    shortcut()
    message.send_keys("BANI PARK: D-92(Z) MeeraMarg Banipark, Jaipur")
    shortcut()
    message.send_keys("Contact: 9784597275 , 9680836271 ")
    shortcut()
    shortcut()
    message.send_keys("*“COOL SUMMER CAMP 2022”*")
    shortcut()
    message.send_keys("SPECIAL PACKAGES FOR KIDS (Age group 3 to 8 years)")
    shortcut()
    message.send_keys("We humbly request to forward this message to your relatives, friends & groups.")
    shortcut()
    # press enter
    keyboard.press(Key.enter)
    keyboard.release(Key.enter)
    # wait for some time
    time.sleep(2)
# Close the browser
driver.quit()
