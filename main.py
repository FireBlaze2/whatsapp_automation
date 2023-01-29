import time
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from pynput.keyboard import Key, Controller
from seleniumbase import BaseCase

filepath = "C:/Users/Raghav/Desktop/contact_list.xlsx"
wb = load_workbook(filepath)
sheet = wb.active
frequency = 113710
# open chrome
driver = webdriver.Chrome('C:\chromedriver\chromedriver')
# open whatsapp homepage for scanning
driver.get("https://web.whatsapp.com")
# wait for some time(in sec)
time.sleep(50)
# open new window
driver.execute_script("window.open('');")
keyboard = Controller()


def shortcut():
    driver.find_element_by_xpath('//*[@title="Type a message"]').send_keys(Keys.SHIFT, Keys.ENTER)


i = 0
l = 0
while i <= frequency:
    while i <= frequency:
        b3 = sheet.cell(row=i+1, column=2)
        # url = "https://wa.me/+91"
        # prints out the link
        print("Sr. no."+(str(i+1)))
        # print(url + str(b3.value))
        # Switch to the new window
        driver.switch_to.window(driver.window_handles[1])
        # open the link
        url = "https://web.whatsapp.com/send?phone=+91" + str(b3.value) + "&text=" + "&app_absent=0"
        driver.get(url)
        # driver.get(url + str(b3.value))
        # # wait for some time(in sec)
        # time.sleep(2)
        # # locating continue to chat
        # continue_to_chat = driver.find_element_by_id("action-button")
        # # continue_to_chat.click()
        # continue_to_chat.click()
        # time.sleep(1)
        # # clicking on use whatsapp web
        # use_whatsapp_web = driver.find_element_by_xpath("// a[contains(text(),'use WhatsApp Web')]")
        # use_whatsapp_web.click()
        # # waiting for the no. to load
        time.sleep(15)
        # attaching image
        try:
            driver.find_element_by_xpath('//div[@title = "Attach"]').click()
        except NoSuchElementException:
            print(str(b3.value) + " NOT ON WHATSAPP")
            l += 1
            i += 1
            break
        image_box = driver.find_element_by_xpath(
            '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')
        # locating image
        image_box.send_keys('C:/Users/Raghav/Desktop/Brochure.jpg')
        time.sleep(2)
        # pressing send button
        driver.find_element_by_xpath('//span[@data-icon="send"]').click()
        # wait for some time
        time.sleep(1)
        # typing message
        message = driver.find_element_by_xpath('//*[@title="Type a message"]')
        message.send_keys("Modern Gurukul Academy @Banipark @Vidyadhar nagar")
        shortcut()
        shortcut()
        message.send_keys("Serving in Banipark since 23 years")
        shortcut()
        shortcut()
        message.send_keys("*Brings to you a one stop centre where your child can get 360° personality development*")
        shortcut()
        message.send_keys("*Join -*")
        shortcut()
        message.send_keys("*• Tuitions (NUR. TO VIII)*")
        shortcut()
        message.send_keys("*• UCMAS*")
        shortcut()
        message.send_keys("*• Dance*")
        shortcut()
        message.send_keys("*• Drawing*")
        shortcut()
        message.send_keys("*• Musical Instruments*")
        shortcut()
        message.send_keys("*• Art and Craft*")
        shortcut()
        message.send_keys("*• Handwriting improvement*")
        shortcut()
        message.send_keys("*• Calligraphy*")
        shortcut()
        message.send_keys("*• Spoken English*")
        shortcut()
        message.send_keys("*• Skating (Banipark)*")
        shortcut()
        shortcut()
        message.send_keys("Let your child learn and explore the world of creativity all under one roof..! All the "
                          "classes will help groom your childs school and co-curricular activities together..")
        shortcut()
        shortcut()
        message.send_keys("We have multiple discount packages available starting from a week to monthly basis.")
        shortcut()
        shortcut()
        message.send_keys("Hurry join now...don't waste even a day!!!")
        shortcut()
        shortcut()
        message.send_keys("BANIPARK: D-92(Z) MeeraMarg Banipark, Jaipur")
        shortcut()
        shortcut()
        message.send_keys("Vidhyadhar Nagar: 144, Opp. MGPS, Shankar Colony Mall Road, Vidhyadhar Nagar, Jaipur")
        shortcut()
        message.send_keys("https://bit.ly/3Q62Emg")
        shortcut()
        message.send_keys("Find us on Instagram -> www.instagram.com/moderngurukulacademy")
        shortcut()
        shortcut()
        message.send_keys("For More Details Contact: 9784597275 , 9680836271")

        # press enter
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
        print("Sent Messages = " + str((i+1)-l))
        i += 1
        # wait for some time
        time.sleep(2)

# Close the browser
driver.quit()
